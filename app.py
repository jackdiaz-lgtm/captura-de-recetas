"""Captura de Recetas — El Taller de Cocina SAS. FastAPI + SQLite."""
import json
import os
import sqlite3
from contextlib import contextmanager
from io import BytesIO

from fastapi import FastAPI, Header, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

DB = os.environ.get("DATABASE_PATH", "captura.db")
PIN = os.environ.get("PIN", "1234")
BASE = os.path.dirname(os.path.abspath(__file__))

app = FastAPI(title="Captura Recetas — El Taller de Cocina")


@contextmanager
def db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys=ON")
    try:
        yield con
        con.commit()
    finally:
        con.close()


def init():
    with db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS canonicos(nombre TEXT PRIMARY KEY, precio_gr REAL);
        CREATE TABLE IF NOT EXISTS recetas(
            id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT UNIQUE NOT NULL,
            area TEXT, tipo TEXT, peso_crudo REAL, rinde_final REAL, porciones REAL,
            precio_carta REAL, tiempo_temp TEXT, notas TEXT, estado TEXT DEFAULT 'PENDIENTE');
        CREATE TABLE IF NOT EXISTS ingredientes(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            receta_id INTEGER REFERENCES recetas(id) ON DELETE CASCADE,
            nombre TEXT, cantidad REAL, nota_nuevo TEXT);
        CREATE TABLE IF NOT EXISTS pasos(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            receta_id INTEGER REFERENCES recetas(id) ON DELETE CASCADE,
            num INTEGER, etapa TEXT, descripcion TEXT, rol TEXT, personas INTEGER,
            min_activo REAL, min_pasivo REAL, equipo TEXT, nota TEXT);
        CREATE TABLE IF NOT EXISTS rendimientos(
            id INTEGER PRIMARY KEY AUTOINCREMENT, ingrediente TEXT,
            bruto REAL, limpio REAL, cocido REAL);
        """)
        # lista de precios: se resincroniza en CADA arranque (upsert), no solo la primera vez
        with open(os.path.join(BASE, "seed_canonicos.json"), encoding="utf-8") as f:
            items = json.load(f)
        c.executemany(
            "INSERT INTO canonicos(nombre,precio_gr) VALUES(?,?) "
            "ON CONFLICT(nombre) DO UPDATE SET precio_gr=excluded.precio_gr", items)
        n = c.execute("SELECT COUNT(*) FROM recetas").fetchone()[0]
        if n == 0:
            with open(os.path.join(BASE, "seed_recetas.json"), encoding="utf-8") as f:
                for r in json.load(f):
                    c.execute("INSERT OR IGNORE INTO recetas(nombre,area,tipo) VALUES(?,?,?)", r)
        # precios oficiales del menu: force=1 siempre pisa; force=0 solo llena vacios
        try:
            with open(os.path.join(BASE, "seed_precios.json"), encoding="utf-8") as f:
                for nombre, precio, force in json.load(f):
                    if force:
                        c.execute("UPDATE recetas SET precio_carta=? WHERE nombre=?", (precio, nombre))
                    else:
                        c.execute("UPDATE recetas SET precio_carta=? WHERE nombre=? AND (precio_carta IS NULL OR precio_carta=0)", (precio, nombre))
        except FileNotFoundError:
            pass


init()


def check(pin):
    if pin != PIN:
        raise HTTPException(401, "PIN incorrecto")


def norm(s):
    import unicodedata
    s = unicodedata.normalize("NFD", (s or "").strip().lower())
    return "".join(ch for ch in s if unicodedata.category(ch) != "Mn")


def costo_gramo_base(con, receta_id, cache, stack):
    """Costo $/gr de una receta-base, resolviendo en cascada:
    ingrediente directo (canonicos) -> otra base capturada (recursivo) -> sin resolver.
    cache: memoiza por receta_id dentro de una misma petición. stack: detecta ciclos."""
    if receta_id in cache:
        return cache[receta_id]
    if receta_id in stack:
        return None  # referencia circular: no resolver
    stack.add(receta_id)
    rec = con.execute("SELECT rinde_final, peso_crudo FROM recetas WHERE id=?", (receta_id,)).fetchone()
    yield_g = (rec["rinde_final"] or rec["peso_crudo"]) if rec else None
    total = 0.0
    completo = yield_g is not None
    for ing in con.execute("SELECT nombre, cantidad FROM ingredientes WHERE receta_id=?", (receta_id,)):
        precio_gr, _ = resolver_ingrediente(con, ing["nombre"], cache, stack)
        if precio_gr is None:
            completo = False
            continue
        total += (ing["cantidad"] or 0) * precio_gr
    stack.discard(receta_id)
    result = (total / yield_g) if (completo and yield_g) else None
    cache[receta_id] = result
    return result


def resolver_ingrediente(con, nombre, cache, stack):
    """Devuelve (precio_gr, origen) para un nombre de ingrediente:
    'directo' si está en Precios_Maestros/canonicos, 'base' si es otra receta
    base ya capturada (resuelto en cascada), None/'pendiente' si no se encuentra."""
    target = norm(nombre)
    for k in con.execute("SELECT nombre, precio_gr FROM canonicos"):
        if norm(k["nombre"]) == target and k["precio_gr"] is not None:
            return k["precio_gr"], "directo"
    for rec in con.execute("SELECT id, nombre FROM recetas"):
        if norm(rec["nombre"]) == target:
            pg = costo_gramo_base(con, rec["id"], cache, stack)
            if pg is not None:
                return pg, "base"
            return None, "base_incompleta"
    return None, "pendiente"


class Receta(BaseModel):
    nombre: str
    area: str = "Restaurante"
    tipo: str = "PLATO"
    peso_crudo: float | None = None
    rinde_final: float | None = None
    porciones: float | None = None
    precio_carta: float | None = None
    tiempo_temp: str | None = None
    notas: str | None = None
    estado: str = "PENDIENTE"


class Ingrediente(BaseModel):
    receta_id: int
    nombre: str
    cantidad: float
    nota_nuevo: str | None = None


class Paso(BaseModel):
    receta_id: int
    num: int
    etapa: str
    descripcion: str
    rol: str | None = None
    personas: int = 1
    min_activo: float | None = None
    min_pasivo: float | None = None
    equipo: str | None = None
    nota: str | None = None


class Rendimiento(BaseModel):
    ingrediente: str
    bruto: float | None = None
    limpio: float | None = None
    cocido: float | None = None


@app.get("/api/login")
def login(x_pin: str = Header("")):
    check(x_pin)
    return {"ok": True}


@app.get("/api/canonicos")
def canonicos(x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        out = [{"nombre": r["nombre"], "precio_gr": r["precio_gr"], "tipo": "ingrediente"}
               for r in c.execute("SELECT * FROM canonicos ORDER BY nombre")]
        # bases ya capturadas: se pueden usar como ingrediente de otra receta (ej. Hogao dentro de un Sancocho)
        out += [{"nombre": r["nombre"], "precio_gr": None, "tipo": "base"}
                for r in c.execute("SELECT nombre FROM recetas WHERE tipo='BASE' ORDER BY nombre")]
        return out


@app.get("/api/recetas")
def recetas(x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        rs = [dict(r) for r in c.execute("SELECT * FROM recetas ORDER BY area, tipo, nombre")]
        cache = {}
        for r in rs:
            r["n_ing"] = c.execute("SELECT COUNT(*) FROM ingredientes WHERE receta_id=?", (r["id"],)).fetchone()[0]
            r["n_pasos"] = c.execute("SELECT COUNT(*) FROM pasos WHERE receta_id=?", (r["id"],)).fetchone()[0]
            total = 0.0
            for ing in c.execute("SELECT nombre, cantidad FROM ingredientes WHERE receta_id=?", (r["id"],)):
                pg, _ = resolver_ingrediente(c, ing["nombre"], cache, {r["id"]})
                if pg is not None:
                    total += (ing["cantidad"] or 0) * pg
            r["costo"] = total
        return rs


@app.post("/api/recetas")
def crear_receta(r: Receta, x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        cur = c.execute(
            "INSERT INTO recetas(nombre,area,tipo,peso_crudo,rinde_final,porciones,precio_carta,tiempo_temp,notas,estado) "
            "VALUES(?,?,?,?,?,?,?,?,?,?) ON CONFLICT(nombre) DO UPDATE SET "
            "area=excluded.area,tipo=excluded.tipo,peso_crudo=excluded.peso_crudo,rinde_final=excluded.rinde_final,"
            "porciones=excluded.porciones,precio_carta=excluded.precio_carta,tiempo_temp=excluded.tiempo_temp,"
            "notas=excluded.notas,estado=excluded.estado",
            (r.nombre, r.area, r.tipo, r.peso_crudo, r.rinde_final, r.porciones,
             r.precio_carta, r.tiempo_temp, r.notas, r.estado))
        rid = c.execute("SELECT id FROM recetas WHERE nombre=?", (r.nombre,)).fetchone()[0]
        return {"id": rid}


@app.get("/api/recetas/{rid}")
def detalle(rid: int, x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        r = c.execute("SELECT * FROM recetas WHERE id=?", (rid,)).fetchone()
        if not r:
            raise HTTPException(404)
        cache = {}
        ings = []
        for i in c.execute("SELECT * FROM ingredientes WHERE receta_id=? ORDER BY id", (rid,)):
            d = dict(i)
            pg, origen = resolver_ingrediente(c, i["nombre"], cache, {rid})
            d["precio_gr"] = pg
            d["origen"] = origen
            d["subtotal"] = round((i["cantidad"] or 0) * pg, 1) if pg is not None else None
            ings.append(d)
        pasos = [dict(p) for p in c.execute("SELECT * FROM pasos WHERE receta_id=? ORDER BY num", (rid,))]
        return {"receta": dict(r), "ingredientes": ings, "pasos": pasos}


@app.delete("/api/recetas/{rid}")
def borrar_receta(rid: int, x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        c.execute("DELETE FROM recetas WHERE id=?", (rid,))
    return {"ok": True}


@app.post("/api/ingredientes")
def add_ing(i: Ingrediente, x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        c.execute("INSERT INTO ingredientes(receta_id,nombre,cantidad,nota_nuevo) VALUES(?,?,?,?)",
                  (i.receta_id, i.nombre, i.cantidad, i.nota_nuevo))
    return {"ok": True}


@app.delete("/api/ingredientes/{iid}")
def del_ing(iid: int, x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        c.execute("DELETE FROM ingredientes WHERE id=?", (iid,))
    return {"ok": True}


@app.post("/api/pasos")
def add_paso(p: Paso, x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        c.execute("INSERT INTO pasos(receta_id,num,etapa,descripcion,rol,personas,min_activo,min_pasivo,equipo,nota) "
                  "VALUES(?,?,?,?,?,?,?,?,?,?)",
                  (p.receta_id, p.num, p.etapa, p.descripcion, p.rol, p.personas,
                   p.min_activo, p.min_pasivo, p.equipo, p.nota))
    return {"ok": True}


@app.delete("/api/pasos/{pid}")
def del_paso(pid: int, x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        c.execute("DELETE FROM pasos WHERE id=?", (pid,))
    return {"ok": True}


@app.get("/api/rendimientos")
def get_rend(x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        return [dict(r) for r in c.execute("SELECT * FROM rendimientos ORDER BY ingrediente")]


@app.post("/api/rendimientos")
def add_rend(r: Rendimiento, x_pin: str = Header("")):
    check(x_pin)
    with db() as c:
        c.execute("INSERT INTO rendimientos(ingrediente,bruto,limpio,cocido) VALUES(?,?,?,?)",
                  (r.ingrediente, r.bruto, r.limpio, r.cocido))
    return {"ok": True}


@app.get("/api/export")
def export(x_pin: str = ""):
    check(x_pin)  # via query param para link de descarga directa
    from openpyxl import Workbook
    wb = Workbook()
    with db() as c:
        ws = wb.active
        ws.title = "RECETAS"
        ws.append(["RECETA", "ÁREA", "TIPO", "PESO CRUDO (g)", "RINDE FINAL (g)", "MERMA %",
                   "PORCIONES", "GRAMAJE/PORCIÓN", "PRECIO CARTA", "COSTO INGREDIENTES",
                   "TIEMPO+TEMP", "NOTAS", "ESTADO"])
        cache_export = {}
        for r in c.execute("SELECT * FROM recetas ORDER BY area, tipo, nombre"):
            costo = 0.0
            for ing in c.execute("SELECT nombre, cantidad FROM ingredientes WHERE receta_id=?", (r["id"],)):
                pg, _ = resolver_ingrediente(c, ing["nombre"], cache_export, {r["id"]})
                if pg is not None:
                    costo += (ing["cantidad"] or 0) * pg
            merma = (1 - r["rinde_final"] / r["peso_crudo"]) if r["peso_crudo"] and r["rinde_final"] else None
            gram = (r["rinde_final"] / r["porciones"]) if r["rinde_final"] and r["porciones"] else None
            ws.append([r["nombre"], r["area"], r["tipo"], r["peso_crudo"], r["rinde_final"],
                       round(merma, 3) if merma is not None else None, r["porciones"],
                       round(gram, 1) if gram else None, r["precio_carta"], round(costo, 1),
                       r["tiempo_temp"], r["notas"], r["estado"]])
        wi = wb.create_sheet("INGREDIENTES")
        wi.append(["RECETA", "INGREDIENTE", "CANTIDAD (g)", "$/gr", "SUBTOTAL", "ORIGEN", "NOTA NUEVO"])
        for i in c.execute(
                "SELECT i.id, re.id AS receta_id, re.nombre AS receta, i.nombre, i.cantidad, i.nota_nuevo "
                "FROM ingredientes i JOIN recetas re ON re.id=i.receta_id ORDER BY re.nombre, i.id"):
            pg, origen = resolver_ingrediente(c, i["nombre"], cache_export, {i["receta_id"]})
            sub = round(i["cantidad"] * pg, 1) if pg is not None else None
            wi.append([i["receta"], i["nombre"], i["cantidad"], pg, sub, origen, i["nota_nuevo"]])
        wp = wb.create_sheet("PROCESOS")
        wp.append(["RECETA", "PASO", "ETAPA", "DESCRIPCIÓN", "ROL", "PERSONAS",
                   "MIN ACTIVO", "MIN PASIVO", "EQUIPO", "NOTA"])
        for p in c.execute("SELECT re.nombre AS receta, p.* FROM pasos p JOIN recetas re ON re.id=p.receta_id "
                           "ORDER BY re.nombre, p.num"):
            wp.append([p["receta"], p["num"], p["etapa"], p["descripcion"], p["rol"],
                       p["personas"], p["min_activo"], p["min_pasivo"], p["equipo"], p["nota"]])
        wr = wb.create_sheet("RENDIMIENTOS")
        wr.append(["INGREDIENTE", "BRUTO (g)", "LIMPIO (g)", "% LIMPIEZA", "COCIDO (g)", "% COCCIÓN"])
        for r in c.execute("SELECT * FROM rendimientos ORDER BY ingrediente"):
            pl = round(r["limpio"] / r["bruto"], 3) if r["bruto"] and r["limpio"] else None
            pc = round(r["cocido"] / r["limpio"], 3) if r["limpio"] and r["cocido"] else None
            wr.append([r["ingrediente"], r["bruto"], r["limpio"], pl, r["cocido"], pc])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Captura_Recetas_Export.xlsx"})


@app.get("/")
def index():
    return FileResponse(os.path.join(BASE, "static", "index.html"))


app.mount("/static", StaticFiles(directory=os.path.join(BASE, "static")), name="static")
